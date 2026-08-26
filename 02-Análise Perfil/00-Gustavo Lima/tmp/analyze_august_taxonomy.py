from __future__ import annotations

import json
import math
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(
    r"C:\Users\Jcluz\AppData\Local\Temp\browser-use\exports\Análise de Crescimento de Autoridade - Gustavo Lima-099d94b8-3077-4b13-86cb-a2001fef5e10.xlsx"
)
OUT = Path(r"C:\Users\Jcluz\chatGPT-Codex\02-Análise Perfil\00-Gustavo Lima\tmp\august_taxonomy_raw.json")


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        return value.replace("�", "â").strip()
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


wb = load_workbook(WORKBOOK, data_only=True, read_only=False)
ws = wb["Postagens 2026"]

rows = []
for row_number in range(1800, 2120):
    values = [clean(ws.cell(row_number, col).value) for col in range(1, 17)]
    if any(value is not None for value in values):
        rows.append({"row": row_number, "values": values})

OUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"rows={len(rows)}")
print(f"saved={OUT}")
for item in rows:
    if 2068 <= item["row"] <= 2080:
        print(item)
