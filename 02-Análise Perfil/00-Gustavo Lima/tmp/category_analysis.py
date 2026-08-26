from __future__ import annotations

import json
import statistics
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\Jcluz\chatGPT-Codex\02-Análise Perfil\00-Gustavo Lima")
WORKBOOK = Path(
    r"C:\Users\Jcluz\AppData\Local\Temp\browser-use\exports\Análise de Crescimento de Autoridade - Gustavo Lima-7f4cc65d-d9de-4a69-b3ca-c383c650e666.xlsx"
)
OUTPUT = ROOT / "tmp" / "august_category_analysis.json"

# First row of each post block in the sheet. The mapping follows the user's
# definition: product/plan education is technical; company, awards and broker
# interviews are authority; family, travel and personal interests are personal.
CATEGORIES = {
    1856: "Técnico",
    1872: "Técnico",
    1880: "Técnico",
    1888: "Autoridade",
    1896: "Técnico",
    1904: "Técnico",
    1912: "Técnico",
    1920: "Autoridade",
    1928: "Técnico",
    1936: "Técnico",
    1944: "Pessoal",
    1952: "Autoridade",
    1960: "Técnico",
    1969: "Técnico",
    1977: "Técnico",
    1985: "Autoridade",
    1993: "Autoridade",
    2001: "Autoridade",
    2009: "Pessoal",
    2017: "Autoridade",
    2025: "Técnico",
    2033: "Autoridade",
    2041: "Técnico",
    2049: "Técnico",
    2057: "Autoridade",
    2065: "Técnico",
    2073: "Autoridade",
    2082: "Técnico",
    2090: "Técnico",
}

METRICS = {
    "views": 5,
    "reach": 6,
    "retention_seconds": 7,
    "hook_rate": 8,
    "interactions": 9,
    "interaction_rate": 10,
    "comments": 11,
    "saves": 13,
    "shares": 14,
    "followers": 15,
}

WEEKDAYS = {
    0: "Segunda",
    1: "Terça",
    2: "Quarta",
    3: "Quinta",
    4: "Sexta",
    5: "Sábado",
    6: "Domingo",
}


def number(value):
    return value if isinstance(value, (int, float)) else None


def summarize(values: list[float]) -> dict[str, float | int | None]:
    if not values:
        return {"n": 0, "sum": 0, "mean": None, "median": None}
    return {
        "n": len(values),
        "sum": round(sum(values), 4),
        "mean": round(statistics.fmean(values), 4),
        "median": round(statistics.median(values), 4),
    }


wb = load_workbook(WORKBOOK, data_only=True, read_only=False)
ws = wb["Postagens 2026"]

posts = []
for row, category in CATEGORIES.items():
    date_value = ws.cell(row, 3).value
    reading_label = str(ws.cell(row, 4).value or "").lower()
    post = {
        "row": row,
        "category": category,
        "title": ws.cell(row, 1).value,
        "format": ws.cell(row, 2).value,
        "date": date_value.isoformat() if isinstance(date_value, datetime) else None,
        "weekday": WEEKDAYS[date_value.weekday()] if isinstance(date_value, datetime) else None,
        "slot": "Dia" if "dia" in reading_label else "Noite" if "noite" in reading_label else None,
    }
    for metric, column in METRICS.items():
        post[metric] = number(ws.cell(row, column).value)
    posts.append(post)

grouped = defaultdict(list)
for post in posts:
    grouped[post["category"]].append(post)

category_stats = {}
for category, category_posts in grouped.items():
    comparable = [post for post in category_posts if post["views"] is not None]
    category_stats[category] = {
        "planned_period_posts": len(category_posts),
        "comparable_24h_posts": len(comparable),
        "metrics": {
            metric: summarize(
                [post[metric] for post in comparable if post[metric] is not None]
            )
            for metric in METRICS
        },
    }


def grouped_metrics(group_name: str) -> dict:
    result = {}
    labels = sorted({post[group_name] for post in posts if post[group_name]})
    if group_name == "weekday":
        labels = list(WEEKDAYS.values())
    for label in labels:
        comparable = [
            post
            for post in posts
            if post[group_name] == label and post["views"] is not None
        ]
        result[label] = {
            "posts": len(comparable),
            "metrics": {
                metric: summarize(
                    [post[metric] for post in comparable if post[metric] is not None]
                )
                for metric in METRICS
            },
        }
    return result

payload = {
    "method": "First 24-hour reading for each post; CPT has no comparable 24-hour metrics.",
    "period_posts": len(posts),
    "comparable_24h_posts": sum(post["views"] is not None for post in posts),
    "mix": {
        category: {
            "posts": len(category_posts),
            "share": round(len(category_posts) / len(posts), 4),
        }
        for category, category_posts in grouped.items()
    },
    "categories": category_stats,
    "timing": {
        "slot": grouped_metrics("slot"),
        "weekday": grouped_metrics("weekday"),
    },
    "posts": posts,
}

OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps(payload, ensure_ascii=False, indent=2))
